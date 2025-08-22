#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
商品信息管理系统导出服务
Windows: 写入模板 → 执行VBA宏 → 导出为xlsx
Mac/Linux: 写入模板 → 保留xlsm，用户打开时宏自动运行
"""

import os
import tempfile
import shutil
import platform
import subprocess
import logging
from datetime import datetime
from io import BytesIO
import openpyxl
import time
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image

# 使用主应用的日志配置
from logging_config import get_logger
logger = get_logger(__name__)

class ExportService:
    """导出服务类"""

    def __init__(self):
        self.template_path = 'templates/product_template.xlsm'

    def export_to_excel(self, products_data, selected_columns):
        """导出商品数据"""
        logger.info(f"\n{'='*50}")
        logger.info(f"🚀 导出服务开始执行")
        logger.info(f"{'='*50}")
        
        temp_files_to_cleanup = []  # 记录需要清理的临时文件
        try:
            logger.info(f"=== 导出开始 ===")
            logger.info(f"输入数据: {len(products_data)} 条记录")
            logger.info(f"选择的列: {selected_columns}")
            logger.info(f"模板路径: {self.template_path}")
            logger.info(f"模板文件存在: {os.path.exists(self.template_path)}")
            
            # 0. 规范化列名（将 image_path 等同于 image）
            logger.info(f"\n📋 步骤1: 规范化列名")
            normalized_columns = self._normalize_columns(selected_columns)
            logger.info(f"规范化后的列: {normalized_columns}")

            # 1. 写入数据到模板
            logger.info(f"\n📝 步骤2: 写入数据到模板")
            temp_template_path = self._write_data_to_template(products_data, normalized_columns)
            temp_files_to_cleanup.append(temp_template_path)
            logger.info(f"临时模板路径: {temp_template_path}")

            # 2. 根据平台执行不同逻辑
            logger.info(f"\n🖥️ 步骤3: 平台检测和导出")
            system_type = platform.system()
            logger.info(f"当前系统: {system_type}")
            
            if system_type == 'Windows':
                logger.info(f"🔧 使用Windows导出逻辑")
                final_excel_data = self._export_windows(temp_template_path)
            else:
                logger.info(f"🍎 使用Mac/Linux导出逻辑")
                final_excel_data = self._export_mac_linux(temp_template_path)

            logger.info(f"\n�� 导出结果")
            logger.info(f"最终数据大小: {len(final_excel_data) if final_excel_data else 0} 字节")
            logger.info(f"=== 导出完成 ===")
            return final_excel_data

        except Exception as e:
            logger.error(f"\n❌ 导出失败: {str(e)}")
            import traceback
            traceback.print_exc()
            return None
        finally:
            logger.info(f"\n🧹 清理临时文件")
            # 清理所有临时文件
            self._cleanup_temp_files(temp_files_to_cleanup)
            logger.info(f"{'='*50}")
            logger.info(f"🏁 导出服务执行结束")
            logger.info(f"{'='*50}\n")

    def _export_windows(self, template_path):
        """Windows系统导出逻辑"""
        try:
            logger.info(f"\n🔧 Windows导出逻辑开始")
            logger.info(f"Windows系统：执行VBA宏美化...")
            logger.info(f"模板路径: {template_path}")
            
            # 调用VBA宏
            logger.info(f"\n📜 步骤3.1: 调用VBA宏")
            self._trigger_vba_macro(template_path)
            logger.info("VBA宏执行完成，开始转换为xlsx...")
            
            # 导出为无宏xlsx
            logger.info(f"\n📊 步骤3.2: 转换为xlsx格式")
            final_excel_data = self._export_to_xlsx_no_macro(template_path)
            logger.info(f"xlsx转换完成，最终数据大小: {len(final_excel_data)} 字节")
            logger.info("✓ Windows导出完成")
            return final_excel_data
            
        except Exception as e:
            logger.error(f"\n❌ Windows导出失败: {str(e)}")
            import traceback
            traceback.print_exc()
            raise

    def _export_mac_linux(self, template_path):
        """Mac/Linux系统导出逻辑"""
        try:
            logger.info("Mac/Linux系统：保留xlsm格式，宏在用户打开时自动执行...")
            # 直接返回xlsm文件内容
            with open(template_path, "rb") as f:
                final_excel_data = f.read()
            logger.info("✓ Mac/Linux导出完成")
            return final_excel_data
        except Exception as e:
            logger.error(f"Mac/Linux导出失败: {str(e)}")
            raise

    def _write_data_to_template(self, products_data, selected_columns):
        """将数据写入模板"""
        try:
            logger.info(f"开始写入模板...")
            temp_dir = tempfile.gettempdir()
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            temp_template_path = os.path.join(temp_dir, f'temp_template_{timestamp}.xlsm')
            logger.info(f"临时模板路径: {temp_template_path}")

            logger.info(f"复制模板文件...")
            shutil.copy2(self.template_path, temp_template_path)
            logger.info(f"模板文件复制完成")
            
            logger.info(f"加载工作簿...")
            workbook = openpyxl.load_workbook(temp_template_path, keep_vba=True)
            logger.info(f"工作簿加载完成，工作表: {workbook.sheetnames}")

            # 找表
            if '商品信息模板' in workbook.sheetnames:
                worksheet = workbook['商品信息模板']
                logger.info(f"找到工作表: 商品信息模板")
            else:
                worksheet = workbook.active
                logger.info(f"使用默认工作表: {worksheet.title}")

            # 清空数据
            logger.info(f"清空现有数据...")
            self._clear_worksheet_data(worksheet)

            # 先清空表头区域（避免模板残留列名）
            # 假设模板表头不超过 20 列
            logger.info(f"清空表头区域...")
            for col in range(1, max(worksheet.max_column, 20) + 1):
                worksheet.cell(row=1, column=col).value = None

            # 删除遗留的 image_path 列（若存在）
            # 遍历当前可见列，若首行等于 image_path 则删除该列
            try:
                col_idx = 1
                while col_idx <= worksheet.max_column:
                    cell_val = worksheet.cell(row=1, column=col_idx).value
                    if isinstance(cell_val, str) and cell_val.strip().lower() == 'image_path':
                        worksheet.delete_cols(col_idx, 1)
                        # 不自增，继续检查当前索引位置（向左移位后的新列）
                        continue
                    col_idx += 1
            except Exception:
                pass

            # 写入表头
            logger.info(f"写入表头...")
            for col_idx, column in enumerate(selected_columns, 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.value = self._get_column_display_name(column)
                self._apply_header_style(cell)
                logger.info(f"表头 {col_idx}: {cell.value}")

            # 写入数据
            logger.info(f"写入数据...")
            for row_idx, product in enumerate(products_data, 2):
                logger.info(f"处理第 {row_idx} 行: {product}")
                for col_idx, column in enumerate(selected_columns, 1):
                    if column == 'image':
                        # 图片列：插入实际图片
                        logger.info(f"处理第{row_idx}行图片列，图片路径: {product.get('image_path', '')}")
                        self._insert_image_to_cell(worksheet, row_idx, col_idx, product.get('image_path', ''))
                        # 设置行高以适应原图（设置更大的行高）
                        worksheet.row_dimensions[row_idx].height = 120
                    else:
                        # 其他列：写入文本值
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.value = self._get_product_value(product, column)
                        self._apply_data_style(cell)
                        logger.info(f"  列 {col_idx} ({column}): {cell.value}")

            # 调整列宽
            logger.info(f"调整列宽...")
            self._adjust_column_widths(worksheet, selected_columns)

            logger.info(f"保存工作簿...")
            workbook.save(temp_template_path)
            workbook.close()
            logger.info(f"工作簿保存完成")

            logger.info(f"✓ 数据已写入模板: {temp_template_path}")
            return temp_template_path
            
        except Exception as e:
            logger.error(f"写入模板失败: {str(e)}")
            import traceback
            traceback.print_exc()
            raise

    def _normalize_columns(self, selected_columns):
        """将来自前端的列名统一成内部标准名。
        - 将 image_path 映射为 image
        - 过滤未知列，保持顺序
        """
        mapping = {
            # 新字段（及兼容旧字段名）
            'doc_date': 'doc_date',
            'customer_name': 'customer_name',
            'product_desc': 'product_desc',
            'unit': 'unit',
            'quantity': 'quantity',
            'unit_price': 'unit_price',
            'unit_discount_rate': 'unit_discount_rate',
            'unit_price_discounted': 'unit_price_discounted',
            'amount': 'amount',
            'image': 'image',
            'remark': 'remark',
            'freight': 'freight',
            'order_discount_rate': 'order_discount_rate',
            'amount_discounted': 'amount_discounted',
            'receivable': 'receivable',
            'paid_total': 'paid_total',
            'balance': 'balance',
            'settlement_account': 'settlement_account',
            'description': 'description',
            'salesperson': 'salesperson',
            'update_time': 'update_time',
            'create_time': 'create_time',
            # 兼容旧键名
            'name': 'customer_name',     # 旧 name 作为客户名称
            'price': 'unit_price',       # 旧 price 作为单价
            'spec': 'unit',              # 旧 spec 作为单位
            'image_path': 'image',
        }
        normalized = []
        for col in selected_columns:
            key = mapping.get(col, None)
            if key and key not in normalized:
                normalized.append(key)
        return normalized

    def _resolve_image_path(self, image_filename: str) -> str:
        """尽量解析图片的绝对路径。"""
        if not image_filename:
            return ''

        # 已是绝对路径
        if os.path.isabs(image_filename) and os.path.exists(image_filename):
            return image_filename

        candidates = []
        # 工程根目录
        project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
        candidates.append(os.path.join(project_root, 'uploads', image_filename))
        # 当前工作目录
        candidates.append(os.path.join(os.getcwd(), 'uploads', image_filename))
        # 直接相对路径
        candidates.append(os.path.join('uploads', image_filename))
        candidates.append(image_filename)

        for p in candidates:
            if os.path.exists(p):
                return p
        return ''

    def _insert_image_to_cell(self, worksheet, row, col, image_path):
        """在指定单元格插入图片"""
        try:
            if not image_path:
                logger.warning(f"图片路径为空")
                return
            # 解析为可用的绝对路径
            full_image_path = self._resolve_image_path(image_path)
            if not full_image_path:
                logger.warning(f"找不到图片文件: {image_path}")
                return
            logger.info(f"正在插入图片: {full_image_path}")
            
            # 直接使用原图，不压缩
            # 将图片插入到Excel
            from openpyxl.drawing.image import Image as XLImage
            excel_img = XLImage(full_image_path)
            
            # 保持原图尺寸，不强制设置宽高
            # 如果需要调整大小，可以在这里设置
            # excel_img.width = 200  # 可以根据需要调整
            # excel_img.height = 150
            
            # 将图片放置在单元格附近
            excel_img.anchor = f'{get_column_letter(col)}{row}'
            
            # 添加图片到工作表
            worksheet.add_image(excel_img)
            
            # 不要在这里删除临时图片文件，让 openpyxl 在保存时处理
            # 我们将在整个导出完成后清理所有临时文件
            logger.info(f"✓ 图片已插入到单元格 {get_column_letter(col)}{row}")
            
        except Exception as e:
            logger.error(f"插入图片失败: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def _cleanup_temp_files(self, temp_files):
        """清理临时文件"""
        for temp_file in temp_files:
            try:
                if os.path.exists(temp_file):
                    # 加入重试，解决 WinError 32 句柄占用
                    for attempt in range(5):
                        try:
                            os.remove(temp_file)
                            logger.info(f"已清理临时文件: {temp_file}")
                            break
                        except PermissionError as pe:
                            logger.warning(f"第{attempt+1}次删除失败(可能句柄占用): {pe}")
                            time.sleep(0.6)
                        except Exception:
                            raise
            except Exception as e:
                logger.error(f"清理临时文件失败 {temp_file}: {str(e)}")
        
        # 清理临时图片文件
        temp_dir = tempfile.gettempdir()
        try:
            for filename in os.listdir(temp_dir):
                if filename.startswith('temp_img_') and filename.endswith('.png'):
                    temp_img_path = os.path.join(temp_dir, filename)
                    os.remove(temp_img_path)
                    logger.info(f"已清理临时图片: {filename}")
        except Exception as e:
            logger.error(f"清理临时图片失败: {str(e)}")

    def _trigger_vba_macro(self, template_path):
        """Windows下触发VBA宏"""
        try:
            logger.info(f"=== VBA宏执行开始 ===")
            logger.info(f"模板路径: {template_path}")
            logger.info(f"模板文件存在: {os.path.exists(template_path)}")
            
            safe_path = template_path.replace("\\", "\\\\")
            logger.info(f"安全路径: {safe_path}")
            
            vbs_script = f'''
Set objExcel = CreateObject("Excel.Application")
objExcel.Visible = False
objExcel.DisplayAlerts = False

Set objWorkbook = objExcel.Workbooks.Open("{safe_path}")

On Error Resume Next
objExcel.Run objWorkbook.Name & "!BeautifySheet"
If Err.Number <> 0 Then
    WScript.Echo "Run macro error: " & Err.Description
    Err.Clear
End If
On Error GoTo 0

' 等待宏执行完成
WScript.Sleep 3000
objWorkbook.Save
objWorkbook.Close False
objExcel.Quit
'''
            logger.info(f"VBS脚本内容:")
            logger.info(vbs_script)

            temp_dir = tempfile.gettempdir()
            vbs_path = os.path.join(temp_dir, f'trigger_macro_{datetime.now().strftime("%Y%m%d_%H%M%S")}.vbs')
            logger.info(f"VBS文件路径: {vbs_path}")

            with open(vbs_path, 'w', encoding='utf-8') as f:
                f.write(vbs_script)
            logger.info(f"VBS文件写入完成")

            logger.info(f"开始执行VBS脚本...")
            # 适当增大超时时间，确保宏有足够时间执行
            result = subprocess.run(['cscript', '//NoLogo', vbs_path], shell=True, timeout=60, capture_output=True, text=True)
            logger.info(f"VBS执行返回码: {result.returncode}")
            logger.info(f"VBS执行输出: {result.stdout}")
            logger.info(f"VBS执行错误: {result.stderr}")
            
            os.remove(vbs_path)
            logger.info(f"VBS文件已删除")

            # 检查执行结果
            if result.returncode != 0 or (result.stdout and "Run macro error" in result.stdout):
                logger.warning("VBA脚本可能未完全执行，已继续导出为xlsx")
            else:
                logger.info("✓ VBA宏执行完成")

        except Exception as e:
            logger.error(f"❌ Windows VBA宏执行失败: {str(e)}")
            import traceback
            traceback.print_exc()

    def _export_to_xlsx_no_macro(self, template_path):
        """导出为不带宏的xlsx文件"""
        try:
            logger.info(f"开始转换为xlsx格式...")
            logger.info(f"输入模板路径: {template_path}")
            logger.info(f"模板文件存在: {os.path.exists(template_path)}")
            
            # 加载工作簿，不保留VBA宏
            workbook = openpyxl.load_workbook(template_path, keep_vba=False)
            logger.info(f"工作簿加载成功，工作表: {workbook.sheetnames}")
            
            # 保存到内存流
            excel_stream = BytesIO()
            workbook.save(excel_stream)
            excel_stream.seek(0)
            excel_data = excel_stream.getvalue()
            excel_stream.close()
            workbook.close()
            
            logger.info(f"✓ 已导出为xlsx格式，数据大小: {len(excel_data)} 字节")
            logger.info(f"✓ xlsx转换完成，文件头: {excel_data[:10]}")
            return excel_data
            
        except Exception as e:
            logger.error(f"xlsx转换失败: {str(e)}")
            import traceback
            traceback.print_exc()
            raise

    def _clear_worksheet_data(self, worksheet):
        for row in range(2, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row, column=col).value = None

    def _get_column_display_name(self, column):
        mapping = {
            'doc_date': '单据日期',
            'customer_name': '客户名称',
            'product_desc': '品名规格',
            'unit': '单位',
            'quantity': '数量',
            'unit_price': '单价',
            'unit_discount_rate': '单价折扣率(%)',
            'unit_price_discounted': '折后单价',
            'amount': '金额',
            'image': '图片',
            'remark': '备注',
            'freight': '运费',
            'order_discount_rate': '整单折扣率(%)',
            'amount_discounted': '折后金额',
            'receivable': '应收款',
            'paid_total': '已收款',
            'balance': '尾款',
            'settlement_account': '结算账户',
            'description': '说明',
            'salesperson': '营业员',
            'update_time': '修改时间',
            'create_time': '创建时间'
        }
        return mapping.get(column, column)

    def _get_product_value(self, product, column):
        try:
            def num(x, default=0.0):
                try:
                    return float(x if x not in (None, '') else default)
                except Exception:
                    return default

            unit_price = num(product.get('unit_price', product.get('price')))
            qty = num(product.get('quantity'))
            unit_rate = num(product.get('unit_discount_rate', 100))
            amount_raw = num(product.get('amount'))
            order_rate = num(product.get('order_discount_rate', 100))
            freight = num(product.get('freight'))
            paid_total = num(product.get('paid_total'))
            # payment_current 字段已移除
            # payment_current 已删除
            if column == 'doc_date':
                return product.get('doc_date') or (product.get('create_time') or '')[:10]
            if column == 'customer_name':
                return product.get('customer_name') or product.get('name', '')
            if column == 'product_desc':
                return product.get('product_desc', '')
            if column == 'unit':
                return product.get('unit') or product.get('spec', '')
            if column == 'quantity':
                return str(int(qty)) if qty.is_integer() else str(qty)
            if column == 'unit_price':
                return f"{unit_price:.2f}"
            if column == 'unit_discount_rate':
                return f"{unit_rate:.2f}"
            if column == 'unit_price_discounted':
                return f"{(unit_price * unit_rate / 100.0):.2f}"
            if column == 'amount':
                base = amount_raw if amount_raw else (unit_price * unit_rate / 100.0 * qty)
                return f"{base:.2f}"
            if column == 'image':
                return ""
            if column == 'remark':
                return product.get('remark', '')
            if column == 'freight':
                return f"{freight:.2f}"
            if column == 'order_discount_rate':
                return f"{order_rate:.2f}"
            if column == 'amount_discounted':
                base = amount_raw if amount_raw else (unit_price * unit_rate / 100.0 * qty)
                return f"{(base * order_rate / 100.0):.2f}"
            if column == 'receivable':
                base = amount_raw if amount_raw else (unit_price * unit_rate / 100.0 * qty)
                discounted = base * order_rate / 100.0
                return f"{(discounted + freight):.2f}"
            if column == 'paid_total':
                return f"{paid_total:.2f}"
            if column == 'balance':
                base = amount_raw if amount_raw else (unit_price * unit_rate / 100.0 * qty)
                discounted = base * order_rate / 100.0
                receivable = discounted + freight
                return f"{(receivable - paid_total):.2f}"
            if column == 'settlement_account':
                return product.get('settlement_account', '')
            if column == 'description':
                return product.get('description', '')
            if column == 'salesperson':
                return product.get('salesperson', '')
            if column == 'update_time':
                return product.get('update_time', '')
            if column == 'create_time':
                return product.get('create_time', '')
            return str(product.get(column, '') or '')
        except Exception:
            return "错误"

    def _apply_header_style(self, cell):
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = border

    def _apply_data_style(self, cell):
        cell.alignment = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = border

    def _adjust_column_widths(self, worksheet, selected_columns):
        widths = {
            'doc_date': 16,
            'customer_name': 18,
            'product_desc': 28,
            'unit': 10,
            'quantity': 12,
            'unit_price': 12,
            'unit_discount_rate': 14,
            'unit_price_discounted': 14,
            'amount': 14,
            'image': 50,
            'remark': 24,
            'freight': 12,
            'order_discount_rate': 14,
            'amount_discounted': 14,
            'receivable': 14,
            'paid_total': 14,
            'balance': 14,
            'settlement_account': 16,
            'description': 20,
            'salesperson': 12,
            'update_time': 20,
            'create_time': 20
        }
        for col_idx, column in enumerate(selected_columns, 1):
            worksheet.column_dimensions[get_column_letter(col_idx)].width = widths.get(column, 15)
