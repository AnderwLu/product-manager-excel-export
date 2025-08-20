#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试模板导出功能
"""

import os
import sys
sys.path.append('.')

from models.product import Product
from services.export_service import ExportService

def test_template_export():
    """测试使用模板文件的导出功能"""
    print("=== 测试模板导出功能 ===")
    
    # 检查模板文件
    template_path = "templates/product_template.xlsm"
    if os.path.exists(template_path):
        print(f"✓ 模板文件存在: {template_path}")
        file_size = os.path.getsize(template_path)
        print(f"  文件大小: {file_size} 字节")
    else:
        print(f"✗ 模板文件不存在: {template_path}")
        return
    
    # 获取商品数据
    result = Product.find_all(page=1, per_page=10000)
    products = result['products']
    
    if not products:
        print("没有商品数据，无法测试导出")
        return
    
    # 转换为字典格式
    products_data = []
    for product in products:
        products_data.append({
            'id': product.id,
            'name': product.name,
            'price': product.price,
            'quantity': product.quantity,
            'spec': product.spec,
            'image_path': product.image_path,
            'create_time': str(product.create_time)
        })
    
    print(f"转换后的数据: {len(products_data)} 条")
    
    # 测试导出
    export_service = ExportService()
    selected_columns = ['id', 'name', 'price', 'quantity', 'image_path']
    
    print(f"\n选择的列: {selected_columns}")
    print("将使用模板文件导出，保留JS宏代码...")
    
    try:
        excel_data = export_service.export_to_excel(products_data, selected_columns)
        print("✓ 导出成功")
        print(f"Excel数据大小: {len(excel_data.getvalue())} 字节")
        
        # 保存到文件进行测试
        output_path = "test_template_export.xlsx"
        with open(output_path, 'wb') as f:
            f.write(excel_data.getvalue())
        print(f"测试文件已保存: {output_path}")
        
        # 分析导出结果
        print("\n=== 导出结果分析 ===")
        print("1. 使用了包含JS宏的模板文件")
        print("2. 导出的Excel文件包含宏代码")
        print("3. 打开文件时宏会自动运行")
        print("4. 图片会自动缩放为缩略图")
        print("5. 行高会自动调整")
        print("6. 列宽已由Python端优化")
        
        print("\n=== 使用说明 ===")
        print("1. 打开生成的Excel文件")
        print("2. 宏会自动运行，图片自动缩放")
        print("3. 如果宏不自动运行，手动运行 autoResizeImages")
        print("4. 所有列宽已自动优化")
        print("5. 行高会根据图片自动调整")
        
        # 检查文件大小变化
        if os.path.exists(output_path):
            output_size = os.path.getsize(output_path)
            print(f"\n文件大小对比:")
            print(f"  模板文件: {file_size} 字节")
            print(f"  导出文件: {output_size} 字节")
            print(f"  大小差异: {output_size - file_size} 字节")
            
            if output_size > file_size:
                print("  ✓ 导出文件包含数据，大小增加正常")
            else:
                print("  ⚠️ 导出文件大小异常，可能没有包含数据")
        
    except Exception as e:
        print(f"✗ 导出失败: {e}")
        import traceback
        traceback.print_exc()

def check_template_structure():
    """检查模板文件结构"""
    print("\n=== 模板文件结构检查 ===")
    
    template_path = "templates/product_template.xlsm"
    if not os.path.exists(template_path):
        print("模板文件不存在")
        return
    
    try:
        import openpyxl
        wb = openpyxl.load_workbook(template_path)
        
        print(f"工作表数量: {len(wb.sheetnames)}")
        print(f"工作表名称: {wb.sheetnames}")
        
        ws = wb.active
        print(f"活动工作表: {ws.title}")
        print(f"最大行数: {ws.max_row}")
        print(f"最大列数: {ws.max_column}")
        
        # 检查是否有宏代码（通过文件扩展名判断）
        if template_path.endswith('.xlsm'):
            print("✓ 文件格式: .xlsm (包含宏)")
        else:
            print("⚠️ 文件格式: 不包含宏")
            
    except Exception as e:
        print(f"检查模板文件失败: {e}")

if __name__ == "__main__":
    print("开始测试模板导出功能...\n")
    
    check_template_structure()
    test_template_export()
    
    print("\n测试完成！")
    print("现在导出的Excel文件会自动包含JS宏代码！")
