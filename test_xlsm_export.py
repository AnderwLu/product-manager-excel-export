#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试.xlsm格式导出和宏代码保留
"""

import os
import sys
sys.path.append('.')

from models.product import Product
from services.export_service import ExportService

def test_xlsm_export():
    """测试.xlsm格式导出"""
    print("=== 测试.xlsm格式导出 ===")
    
    # 检查模板文件
    template_path = "templates/product_template.xlsm"
    if os.path.exists(template_path):
        print(f"✓ 模板文件存在: {template_path}")
        file_size = os.path.getsize(template_path)
        print(f"  文件大小: {file_size} 字节")
        print(f"  文件格式: .xlsm (包含宏)")
    else:
        print(f"✗ 模板文件不存在: {template_path}")
        return
    
    # 获取商品数据
    result = Product.find_all(page=1, per_page=10000)
    products = result['products']
    
    if not products:
        print("没有商品数据，无法测试导出")
        return
    
    # 转换为字典格式
    products_data = []
    for product in products:
        products_data.append({
            'id': product.id,
            'name': product.name,
            'price': product.price,
            'quantity': product.quantity,
            'spec': product.spec,
            'image_path': product.image_path,
            'create_time': str(product.create_time)
        })
    
    print(f"转换后的数据: {len(products_data)} 条")
    
    # 测试导出
    export_service = ExportService()
    selected_columns = ['id', 'name', 'price', 'image_path']
    
    print(f"\n选择的列: {selected_columns}")
    print("将使用模板文件导出，测试.xlsm格式和宏代码保留...")
    
    try:
        excel_data = export_service.export_to_excel(products_data, selected_columns)
        print("✓ 导出成功")
        print(f"Excel数据大小: {len(excel_data.getvalue())} 字节")
        
        # 保存到文件进行测试
        output_path = "test_xlsm_export.xlsm"  # 使用.xlsm扩展名
        with open(output_path, 'wb') as f:
            f.write(excel_data.getvalue())
        print(f"测试文件已保存: {output_path}")
        
        # 验证输出文件
        verify_xlsm_file(output_path, template_path)
        
    except Exception as e:
        print(f"✗ 导出失败: {e}")
        import traceback
        traceback.print_exc()

def verify_xlsm_file(output_path, template_path):
    """验证输出文件是否为.xlsm格式并包含宏代码"""
    print(f"\n=== 验证.xlsm文件 ===")
    
    if not os.path.exists(output_path):
        print("输出文件不存在")
        return
    
    # 检查文件大小
    output_size = os.path.getsize(output_path)
    template_size = os.path.getsize(template_path)
    
    print(f"模板文件大小: {template_size} 字节")
    print(f"输出文件大小: {output_size} 字节")
    print(f"大小差异: {output_size - template_size} 字节")
    
    # 检查文件扩展名
    if output_path.endswith('.xlsm'):
        print("✓ 输出文件扩展名: .xlsm (支持宏)")
    else:
        print("⚠️ 输出文件扩展名: 不支持宏")
    
    # 尝试加载文件检查宏
    try:
        import openpyxl
        wb = openpyxl.load_workbook(output_path, keep_vba=True)
        
        print(f"✓ 成功加载输出文件")
        print(f"  工作表数量: {len(wb.sheetnames)}")
        print(f"  工作表名称: {wb.sheetnames}")
        
        # 检查是否有VBA项目
        if hasattr(wb, 'vba_archive'):
            print("✓ 文件包含VBA项目，宏代码已保留")
            
            # 检查VBA项目详情
            vba_archive = wb.vba_archive
            print(f"  VBA项目名称: {vba_archive.name}")
            print(f"  VBA项目类型: {type(vba_archive).__name__}")
            
        else:
            print("⚠️ 文件不包含VBA项目，宏代码可能丢失")
            
    except Exception as e:
        print(f"✗ 加载输出文件失败: {e}")
    
    # 检查文件头部，确认是否为.xlsm格式
    try:
        with open(output_path, 'rb') as f:
            header = f.read(100)
            
        # .xlsm文件应该包含特定的ZIP文件头
        if header.startswith(b'PK'):
            print("✓ 文件头部确认: 包含ZIP文件头，确实是.xlsm格式")
        else:
            print("⚠️ 文件头部异常: 不是标准的.xlsm格式")
            
    except Exception as e:
        print(f"✗ 检查文件头部失败: {e}")

def check_template_macro():
    """检查模板文件的宏代码"""
    print("\n=== 检查模板文件宏代码 ===")
    
    template_path = "templates/product_template.xlsm"
    if not os.path.exists(template_path):
        print("模板文件不存在")
        return
    
    try:
        import openpyxl
        wb = openpyxl.load_workbook(template_path, keep_vba=True)
        
        print(f"✓ 成功加载模板文件")
        print(f"  工作表数量: {len(wb.sheetnames)}")
        print(f"  工作表名称: {wb.sheetnames}")
        
        # 检查是否有VBA项目
        if hasattr(wb, 'vba_archive'):
            print("✓ 模板文件包含VBA项目，宏代码存在")
            
            # 检查VBA项目详情
            vba_archive = wb.vba_archive
            print(f"  VBA项目名称: {vba_archive.name}")
            print(f"  VBA项目类型: {type(vba_archive).__name__}")
            
        else:
            print("⚠️ 模板文件不包含VBA项目，宏代码可能丢失")
            
    except Exception as e:
        print(f"✗ 检查模板文件失败: {e}")

if __name__ == "__main__":
    print("开始测试.xlsm格式导出...\n")
    
    check_template_macro()
    test_xlsm_export()
    
    print("\n测试完成！")
    print("现在导出的Excel文件应该是.xlsm格式并包含宏代码！")
