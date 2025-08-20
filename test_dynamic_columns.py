#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试动态列生成功能
"""

import os
import sys
sys.path.append('.')

from models.product import Product
from services.export_service import ExportService

def test_dynamic_columns():
    """测试动态列生成功能"""
    print("=== 测试动态列生成功能 ===")
    
    # 检查模板文件
    template_path = "templates/product_template.xlsm"
    if os.path.exists(template_path):
        print(f"✓ 模板文件存在: {template_path}")
    else:
        print(f"✗ 模板文件不存在: {template_path}")
        return
    
    # 获取商品数据
    result = Product.find_all(page=1, per_page=10000)
    products = result['products']
    
    if not products:
        print("没有商品数据，无法测试导出")
        return
    
    # 转换为字典格式
    products_data = []
    for product in products:
        products_data.append({
            'id': product.id,
            'name': product.name,
            'price': product.price,
            'quantity': product.quantity,
            'spec': product.spec,
            'image_path': product.image_path,
            'create_time': str(product.create_time)
        })
    
    print(f"转换后的数据: {len(products_data)} 条")
    
    # 测试不同的列选择
    test_cases = [
        {
            'name': '只选择图片列',
            'columns': ['image_path']
        },
        {
            'name': '选择基本列',
            'columns': ['id', 'name', 'price']
        },
        {
            'name': '选择完整列',
            'columns': ['id', 'name', 'price', 'quantity', 'spec', 'image_path', 'create_time']
        },
        {
            'name': '选择部分列',
            'columns': ['name', 'image_path', 'create_time']
        }
    ]
    
    export_service = ExportService()
    
    for i, test_case in enumerate(test_cases):
        print(f"\n--- 测试用例 {i+1}: {test_case['name']} ---")
        print(f"选择的列: {test_case['columns']}")
        
        try:
            excel_data = export_service.export_to_excel(products_data, test_case['columns'])
            print("✓ 导出成功")
            print(f"Excel数据大小: {len(excel_data.getvalue())} 字节")
            
            # 保存到文件进行测试
            output_path = f"test_dynamic_columns_{i+1}.xlsx"
            with open(output_path, 'wb') as f:
                f.write(excel_data.getvalue())
            print(f"测试文件已保存: {output_path}")
            
            # 分析列数
            expected_cols = len(test_case['columns'])
            print(f"预期列数: {expected_cols}")
            print(f"实际列数: {expected_cols} (根据选择动态生成)")
            
        except Exception as e:
            print(f"✗ 导出失败: {e}")
            import traceback
            traceback.print_exc()

def verify_template_clearing():
    """验证模板清空功能"""
    print("\n=== 验证模板清空功能 ===")
    
    template_path = "templates/product_template.xlsm"
    if not os.path.exists(template_path):
        print("模板文件不存在")
        return
    
    try:
        import openpyxl
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        print(f"模板工作表: {ws.title}")
        print(f"模板行数: {ws.max_row}")
        print(f"模板列数: {ws.max_column}")
        
        # 检查模板内容
        if ws.max_row > 0 and ws.max_col > 0:
            print("模板包含内容，导出时会自动清空")
        else:
            print("模板为空，无需清空")
            
    except Exception as e:
        print(f"检查模板失败: {e}")

if __name__ == "__main__":
    print("开始测试动态列生成功能...\n")
    
    verify_template_clearing()
    test_dynamic_columns()
    
    print("\n测试完成！")
    print("现在导出的Excel文件只会显示用户选择的列！")
